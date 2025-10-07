"""
Tests for ingestion_standardization module.

Uses pytest with tmp_path fixtures for file I/O; no external network calls.
"""

from pathlib import Path

import pandas as pd
import pytest

from src.ingestion_standardization import (
    DEFAULT_STATE_POP,
    Paths,
    _parse_nhtsa_date,
    _safe_month_parse,
    ingest_afdc_historical,
    ingest_eia_price,
    ingest_eia_sales,
    ingest_recalls,
    run_all,
    standardize_afdc,
    standardize_eia,
    standardize_recalls,
)


# ---------------------------------------------------------------------------
# Helper function tests
# ---------------------------------------------------------------------------
def test_safe_month_parse():
    """Test robust month parsing."""
    assert _safe_month_parse("2024-01") == pd.Timestamp("2024-01-01")
    assert _safe_month_parse("Jan-2024") == pd.Timestamp("2024-01-01")
    assert _safe_month_parse("1/2024") == pd.Timestamp("2024-01-01")
    assert _safe_month_parse("2024 M01") == pd.Timestamp("2024-01-01")
    assert _safe_month_parse(None) is None
    assert _safe_month_parse("") is None


def test_parse_nhtsa_date():
    """Test NHTSA ID date parsing."""
    assert _parse_nhtsa_date("20E003") == pd.Timestamp("2020-01-01")
    assert _parse_nhtsa_date("23V456") == pd.Timestamp("2023-01-01")
    assert _parse_nhtsa_date(None) is None
    assert _parse_nhtsa_date("invalid") is None


# ---------------------------------------------------------------------------
# AFDC tests
# ---------------------------------------------------------------------------
def test_ingest_afdc_historical_empty(tmp_path):
    """Test AFDC ingestion with no files."""
    result = ingest_afdc_historical([tmp_path / "nonexistent.csv"])
    assert result.empty


def test_ingest_afdc_historical(tmp_path):
    """Test AFDC ingestion with valid CSV."""
    csv_path = tmp_path / "alt_fuel_stations_historical_day (Oct 7 2021).csv"
    csv_path.write_text(
        "Fuel Type Code,State,EV DC Fast Count,EV Level2 EVSE Num\n"
        "ELEC,CA,5,10\n"
        "ELEC,NY,2,8\n"
    )

    result = ingest_afdc_historical([csv_path])
    assert len(result) == 2
    assert "snapshot_date" in result.columns
    assert result["snapshot_date"].iloc[0] == pd.Timestamp("2021-10-07")


def test_ingest_afdc_historical_fallback_connectors(tmp_path):
    """Test AFDC ingestion with connector fallback."""
    csv_path = tmp_path / "stations.csv"
    csv_path.write_text(
        "Fuel Type Code,State,EV Connector Types\n"
        "ELEC,CA,J1772; CCS; CHAdeMO\n"
        "ELEC,NY,J1772\n"
    )

    result = ingest_afdc_historical([csv_path])
    assert len(result) == 2


def test_standardize_afdc_empty():
    """Test AFDC standardization with empty input."""
    ports, dcfast = standardize_afdc(pd.DataFrame())
    assert ports.empty
    assert dcfast.empty


def test_standardize_afdc(tmp_path):
    """Test AFDC standardization."""
    # Create sample data
    df = pd.DataFrame(
        {
            "Fuel Type Code": ["ELEC", "ELEC", "ELEC"],
            "State": ["CA", "CA", "NY"],
            "EV DC Fast Count": [5, 3, 2],
            "EV Level2 EVSE Num": [10, 8, 15],
            "EV Level1 EVSE Num": [0, 0, 0],
            "snapshot_date": ["2021-10-07", "2021-10-07", "2021-10-07"],
        }
    )

    ports, dcfast = standardize_afdc(df, state_pop=DEFAULT_STATE_POP)

    # Check outputs
    assert not ports.empty
    assert not dcfast.empty
    assert "ports_per_100k" in ports.columns
    assert "dcfast_share" in dcfast.columns
    assert ports["state"].isin(["CA", "NY"]).all()


def test_standardize_afdc_invalid_states(tmp_path):
    """Test AFDC standardization filters invalid states."""
    df = pd.DataFrame(
        {
            "Fuel Type Code": ["ELEC", "ELEC"],
            "State": ["CA", "XX"],  # XX is invalid
            "EV DC Fast Count": [5, 3],
            "EV Level2 EVSE Num": [10, 8],
            "snapshot_date": ["2021-10-07", "2021-10-07"],
        }
    )

    ports, dcfast = standardize_afdc(df)
    assert len(ports) == 1  # Only CA should remain
    assert ports["state"].iloc[0] == "CA"


# ---------------------------------------------------------------------------
# EIA tests
# ---------------------------------------------------------------------------
def test_ingest_eia_price_missing_file(tmp_path):
    """Test EIA price ingestion with missing file."""
    result = ingest_eia_price(tmp_path / "nonexistent.xlsx")
    assert result.empty


def test_ingest_eia_price(tmp_path):
    """Test EIA price ingestion with valid Excel file."""
    xlsx_path = tmp_path / "table_5_06_a.xlsx"

    # Create a mock Excel file with multi-level headers matching real EIA format
    # Use openpyxl to write multi-level headers properly
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active

    # Row 1-2: metadata (skip)
    ws.append(["Table 5.06.A"])
    ws.append(["Average Retail Price of Electricity"])

    # Row 3: First level headers (sectors)
    ws.append(["", "Residential", "Residential", "Commercial"])

    # Row 4: Second level headers (dates)
    ws.append(["State", "2024-01", "2024-02", "2024-01"])

    # Data rows
    ws.append(["California", 19.5, 19.8, 18.0])
    ws.append(["New York", 21.3, 21.5, 19.0])
    ws.append(["Texas", 13.2, 13.4, 12.0])

    wb.save(xlsx_path)

    result = ingest_eia_price(xlsx_path)
    assert not result.empty
    assert "state" in result.columns
    assert "month" in result.columns
    assert "cents_per_kWh" in result.columns


def test_ingest_eia_sales(tmp_path):
    """Test EIA sales ingestion with valid Excel file."""
    xlsx_path = tmp_path / "table_5_04_b.xlsx"

    # Create a mock Excel file with multi-level headers matching real EIA format
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active

    # Row 1-2: metadata (skip)
    ws.append(["Table 5.04.B"])
    ws.append(["Retail Sales of Electricity"])

    # Row 3: First level headers (sectors)
    ws.append(["", "Residential", "Residential"])

    # Row 4: Second level headers (dates)
    ws.append(["State", "2024-01", "2024-02"])

    # Data rows
    ws.append(["California", 1000, 1100])
    ws.append(["New York", 800, 850])

    wb.save(xlsx_path)

    result = ingest_eia_sales(xlsx_path)
    assert not result.empty
    assert "state" in result.columns
    assert "month" in result.columns
    assert "MWh" in result.columns


def test_standardize_eia():
    """Test EIA standardization."""
    price = pd.DataFrame(
        {
            "state": ["CA", "NY", "CA"],
            "month": pd.to_datetime(["2024-01", "2024-01", "2024-02"]),
            "cents_per_kWh": [19.5, 21.3, 19.8],
        }
    )
    sales = pd.DataFrame(
        {
            "state": ["CA", "NY", "CA"],
            "month": pd.to_datetime(["2024-01", "2024-01", "2024-02"]),
            "MWh": [1000, 800, 1100],
        }
    )

    price_out, sales_out = standardize_eia(price, sales)
    assert not price_out.empty
    assert not sales_out.empty
    assert price_out["state"].is_monotonic_increasing or price_out[
        "month"
    ].is_monotonic_increasing


# ---------------------------------------------------------------------------
# Recalls tests
# ---------------------------------------------------------------------------
def test_ingest_recalls_empty(tmp_path):
    """Test recalls ingestion with no files."""
    result = ingest_recalls([tmp_path / "nonexistent.csv"])
    assert result.empty


def test_ingest_recalls(tmp_path):
    """Test recalls ingestion with valid CSV."""
    csv_path = tmp_path / "recalls.csv"
    csv_path.write_text(
        '"NHTSA ID","MAKE","MODEL","MODEL YEAR"\n'
        '"20E003","RIVIAN","R1T","2022"\n'
        '"21V456","TESLA","MODEL 3","2021"\n'
    )

    result = ingest_recalls([csv_path])
    assert len(result) == 2
    assert "NHTSA ID" in result.columns or "NHTSA_ID" in result.columns


def test_standardize_recalls_empty():
    """Test recalls standardization with empty input."""
    result = standardize_recalls(pd.DataFrame())
    assert result.empty


def test_standardize_recalls(tmp_path):
    """Test recalls standardization."""
    df = pd.DataFrame(
        {
            "NHTSA ID": ["20E003", "20E004", "21V456"],
            "MAKE": ["RIVIAN", "RIVIAN", "TESLA"],
            "MODEL": ["R1T", "R1S", "MODEL 3"],
            "MODEL YEAR": [2022, 2022, 2021],
        }
    )

    result = standardize_recalls(df, makes=["RIVIAN", "TESLA"])
    assert not result.empty
    assert "make" in result.columns
    assert set(result["make"].unique()).issubset({"RIVIAN", "TESLA"})


def test_standardize_recalls_filter_makes():
    """Test recalls filter by make."""
    df = pd.DataFrame(
        {
            "NHTSA ID": ["20E003", "20E004", "21V456"],
            "MAKE": ["RIVIAN", "FORD", "TESLA"],
            "MODEL": ["R1T", "F-150", "MODEL 3"],
        }
    )

    result = standardize_recalls(df, makes=["RIVIAN"])
    assert len(result) > 0
    assert (result["make"] == "RIVIAN").all()


# ---------------------------------------------------------------------------
# Integration tests
# ---------------------------------------------------------------------------
def test_run_all_with_minimal_data(tmp_path):
    """Test full pipeline with minimal valid data."""
    # Create minimal AFDC file
    afdc_path = tmp_path / "afdc.csv"
    afdc_path.write_text(
        "Fuel Type Code,State,EV DC Fast Count,EV Level2 EVSE Num\n"
        "ELEC,CA,5,10\n"
    )

    # Create minimal EIA price file
    import openpyxl

    price_path = tmp_path / "price.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Table 5.06.A"])
    ws.append(["Average Retail Price"])
    ws.append(["", "Residential"])
    ws.append(["State", "2024-01"])
    ws.append(["California", 19.5])
    wb.save(price_path)

    # Create minimal EIA sales file
    sales_path = tmp_path / "sales.xlsx"
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.append(["Table 5.04.B"])
    ws2.append(["Retail Sales"])
    ws2.append(["", "Residential"])
    ws2.append(["State", "2024-01"])
    ws2.append(["California", 1000])
    wb2.save(sales_path)

    # Create minimal recalls file
    recalls_path = tmp_path / "recalls.csv"
    recalls_path.write_text(
        '"NHTSA ID","MAKE"\n' '"20E003","RIVIAN"\n'
    )

    # Create Paths object
    paths = Paths(
        afdc_files=[afdc_path],
        eia_price_xlsx=price_path,
        eia_sales_xlsx=sales_path,
        recall_files=[recalls_path],
    )

    # Run pipeline
    out_dir = tmp_path / "output"
    run_all(paths, out_dir)

    # Check outputs exist
    assert (out_dir / "afdc_ports_per_100k_by_state_year.csv").exists()
    assert (out_dir / "afdc_dcfast_share_by_state_year.csv").exists()
    assert (out_dir / "eia_res_price_residential_state_month.csv").exists()
    assert (out_dir / "eia_res_sales_per_customer_state_month.csv").exists()

    # Check master panel
    if (out_dir / "master_panel_state_month.csv").exists():
        master = pd.read_csv(out_dir / "master_panel_state_month.csv")
        assert not master.empty


def test_run_all_missing_files(tmp_path):
    """Test pipeline handles missing files gracefully."""
    paths = Paths(
        afdc_files=[tmp_path / "missing.csv"],
        eia_price_xlsx=tmp_path / "missing.xlsx",
        eia_sales_xlsx=tmp_path / "missing.xlsx",
        recall_files=[tmp_path / "missing.csv"],
    )

    out_dir = tmp_path / "output"
    # Should not raise; should log warnings
    run_all(paths, out_dir)
    assert out_dir.exists()


def test_state_population_coverage():
    """Test that DEFAULT_STATE_POP covers all valid state codes."""
    from src.ingestion_standardization import STATE_ABBR

    pop_states = set(DEFAULT_STATE_POP["state"])
    assert STATE_ABBR.issubset(pop_states)


def test_state_name_mapping_coverage():
    """Test state name to abbreviation mapping."""
    from src.ingestion_standardization import STATE_NAME_TO_ABBR

    # Check some key states
    assert STATE_NAME_TO_ABBR["California"] == "CA"
    assert STATE_NAME_TO_ABBR["New York"] == "NY"
    assert STATE_NAME_TO_ABBR["District of Columbia"] == "DC"

